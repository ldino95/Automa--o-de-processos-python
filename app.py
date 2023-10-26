from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl

produto_item = ("Console PS5")

#Entrar no site da https://www.pelando.com.br/
driver = webdriver.Chrome()
driver.maximize_window()
driver.get ('https://www.pelando.com.br/')
sleep (5)

#Fazer digitar do produto "PS5"
selecionar_input = driver.find_element(By.XPATH,"//div[@class='sc-gYbzsP ihylnh sc-eJjYHh hgrehZ']")
selecionar_input.click()
sleep(5)
input_pesqusisar = driver.find_element(By.XPATH,"//input[@id='suggestions']")
input_pesqusisar.send_keys(produto_item)
sleep(5)

#Clicar na lupa 
botao_lupa = driver.find_element(By.XPATH,"//i[@title='Buscar']")
botao_lupa.click()
sleep(10)

#Pegar 5 primeiro da lista [Nome, Valor, Vendedor e Data] 
nome_anuncios = driver.find_elements(By.XPATH,"(//a[@class='sc-khsqcC iJiOxL'])")
lista_nome_anuncios = []
for anuncio in nome_anuncios[:5]:  # Pegar os 5 primeiros
    lista_nome_anuncios.append(anuncio.text)

valor_anuncio = driver.find_elements(By.XPATH,"//div[@class='sc-iAEawV gDfGVi sc-iOeugr gZuqIb']")
lista_valor_anuncio = []
for valor in valor_anuncio[:5]:  # Pegar os 5 primeiros
    lista_valor_anuncio.append(valor.text)

vendedor_anuncio = driver.find_elements(By.XPATH,"//a[@class='sc-gScZFl gbYuDt']")
lista_vendedor_anuncio = []
for vendedor in vendedor_anuncio[:5]:  # Pegar os 5 primeiros
    lista_vendedor_anuncio.append(vendedor.text)

data_anuncio = driver.find_elements(By.XPATH,"//div[@title='Hora da publicação']")
lista_data_anuncio = []
for data in data_anuncio[:5]:  # Pegar os 5 primeiros
    lista_data_anuncio.append(data.text)

#Guardar tudo no excel.
workbook = openpyxl.load_workbook('dados.xlsx')
print(workbook.sheetnames)

try:
    #código para inserir dados em uma página existente
    #acessar pagina dos dados
    pagina_anuncio = workbook['Produto ' + str(produto_item)]
    #criar nome das colunas
    pagina_anuncio ['A1'].value = "Nome" 
    pagina_anuncio ['B1'].value = "Valor" 
    pagina_anuncio ['C1'].value = "Vendedor" 
    pagina_anuncio ['D1'].value = "Data"
    #adicionar [Nome] 
    for i, nome in enumerate(lista_nome_anuncios, start=2):
        pagina_anuncio.cell(row=i, column=1, value=nome)
    #adicionar [Valor] 
    for i, valor in enumerate(lista_valor_anuncio, start=2):
        pagina_anuncio.cell(row=i, column=2, value=valor)
    #adicionar [Vendedor] 
    for i, vendedor in enumerate(lista_vendedor_anuncio, start=2):
        pagina_anuncio.cell(row=i, column=3, value=vendedor)
    #adicionar [Data] 
    for i, data in enumerate(lista_data_anuncio, start=2):
        pagina_anuncio.cell(row=i, column=4, value=data)

    workbook.save('Dados.xlsx')
    driver.close()
    sleep(5)

except Exception as error:
    # código para criar uma página do zero e inserir as informações
    workbook.create_sheet('Produto ' + str(produto_item))
    print(workbook.sheetnames)
    pagina_anuncio = workbook['Produto ' + str(produto_item)]
    #criar nome das colunas
    pagina_anuncio ['A1'].value = "Nome" 
    pagina_anuncio ['B1'].value = "Valor" 
    pagina_anuncio ['C1'].value = "Vendedor" 
    pagina_anuncio ['D1'].value = "Data"
    #adicionar [Nome] 
    for i, nome in enumerate(lista_nome_anuncios, start=2):
        pagina_anuncio.cell(row=i, column=1, value=nome)
    #adicionar [Valor] 
    for i, valor in enumerate(lista_valor_anuncio, start=2):
        pagina_anuncio.cell(row=i, column=2, value=valor)
    #adicionar [Vendedor] 
    for i, vendedor in enumerate(lista_vendedor_anuncio, start=2):
        pagina_anuncio.cell(row=i, column=3, value=vendedor)
    #adicionar [Data] 
    for i, data in enumerate(lista_data_anuncio, start=2):
        pagina_anuncio.cell(row=i, column=4, value=data)

    workbook.save('Dados.xlsx')
    driver.close()
    sleep(5)
